"""
Dify 批量下载后处理：将三列 CSV 展开为格式化的验收 Excel。

输出 local/ 目录（已 gitignore），不会被提交到 GitHub。

用法：
    python flatten_batch_output.py <dify_download.csv> [输出文件名]

输出列：
    候选人别名 | 处理日期 | 综合匹配分 | 建议 | 技能匹配 | 相关经验 |
    项目相关性 | 综合质量 | 匹配技能 | 待补充信息 | 风险标记 |
    解析状态 | 人工复核结论 | 备注
"""

import csv
import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = Path(__file__).parent / "local"

# ── 样式常量（对齐 candidate_review_template.xlsx） ──────────────────────
TEAL_FILL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
WHITE_BOLD_16 = Font(bold=True, color="FFFFFF", size=16)
HEADER_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
HEADER_FONT = Font(bold=True, color="134E4A")
THIN_GREEN = Side(style="thin", color="99B9B4")
THIN_GRAY = Side(style="thin", color="D1D5DB")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_WRAP = Alignment(vertical="top", wrap_text=True)

FIELDS = [
    ("候选人别名", 15),
    ("处理日期", 13),
    ("综合匹配分", 12),
    ("建议", 16),
    ("技能匹配", 12),
    ("相关经验", 12),
    ("项目相关性", 13),
    ("综合质量", 12),
    ("匹配技能", 25),
    ("待补充信息", 28),
    ("风险标记", 28),
    ("解析状态", 12),
    ("人工复核结论", 22),
    ("备注", 26),
]


def _join(v):
    if isinstance(v, list):
        return "、".join(str(x) for x in v)
    return v


def flatten(input_path: str, output_name: str = None):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if output_name is None:
        output_name = f"{Path(input_path).stem}_验收就绪.xlsx"
    output_path = OUTPUT_DIR / output_name

    # ── 解析 Dify 下载 CSV ──────────────────────────────────────────
    rows = []
    with open(input_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            raw = row.get("生成结果", "")
            try:
                result = json.loads(raw)
            except (json.JSONDecodeError, TypeError):
                result = {}

            if "parsed_json" in result and isinstance(result["parsed_json"], str):
                try:
                    inner = json.loads(result["parsed_json"])
                    if isinstance(inner, dict) and "match_score" in inner:
                        result = {**inner, **result}
                except (json.JSONDecodeError, TypeError):
                    pass

            rows.append({
                "候选人别名": result.get("candidate_name", ""),
                "处理日期": date.today().isoformat(),
                "综合匹配分": result.get("match_score", ""),
                "建议": result.get("recommendation", ""),
                "技能匹配": result.get("skill_match", ""),
                "相关经验": result.get("experience_relevance", ""),
                "项目相关性": result.get("project_relevance", ""),
                "综合质量": result.get("overall_quality", ""),
                "匹配技能": _join(result.get("matched_skills", "")),
                "待补充信息": _join(result.get("missing_information", "")),
                "风险标记": _join(result.get("risk_flags", "")),
                "解析状态": result.get("parse_status", ""),
                "人工复核结论": "",
                "备注": "",
            })

    if not rows:
        print("WARN: No rows extracted")
        return

    # ── 构建 Excel ──────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "候选人汇总"
    ws.showGridLines = False

    ncols = len(FIELDS)
    nrows = len(rows)

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, "AI 简历初筛与岗位匹配助手｜人工复核汇总")
    ws.cell(1, 1).fill = TEAL_FILL
    ws.cell(1, 1).font = WHITE_BOLD_16
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # 表头行
    header_row = 3
    for ci, (name, width) in enumerate(FIELDS, 1):
        cell = ws.cell(header_row, ci, name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = Border(top=THIN_GREEN, bottom=THIN_GREEN,
                             left=THIN_GREEN, right=THIN_GREEN)

    # 数据行
    for ri, row in enumerate(rows, header_row + 1):
        for ci, (name, _) in enumerate(FIELDS, 1):
            val = row[name]
            cell = ws.cell(ri, ci, val)
            cell.alignment = TOP_WRAP
            cell.border = Border(top=THIN_GRAY, bottom=THIN_GRAY,
                                 left=THIN_GRAY, right=THIN_GRAY)

    # 列宽
    for ci, (_, width) in enumerate(FIELDS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    # 行高
    for ri in range(header_row + 1, header_row + nrows + 1):
        ws.row_dimensions[ri].height = 48

    # 条件格式：综合匹配分列（第3列）颜色渐变
    score_col_letter = get_column_letter(3)
    score_range = f"{score_col_letter}{header_row + 1}:{score_col_letter}{header_row + nrows}"
    ws.conditional_formatting.add(score_range,
        ColorScaleRule(start_type="min", start_color="FECACA",
                       mid_type="percentile", mid_value=50, mid_color="FEF3C7",
                       end_type="max", end_color="BBF7D0"))

    # PASS/FAIL 条件格式：解析状态列
    status_col_letter = get_column_letter(12)
    status_range = f"{status_col_letter}{header_row + 1}:{status_col_letter}{header_row + nrows}"
    ws.conditional_formatting.add(status_range,
        CellIsRule(operator="equal", formula=['"success"'],
                   fill=PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
                   font=Font(color="166534", bold=True)))
    ws.conditional_formatting.add(status_range,
        CellIsRule(operator="notEqual", formula=['"success"'],
                   fill=PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
                   font=Font(color="991B1B", bold=True)))

    # 数据验证：建议列
    dv_rec = DataValidation(type="list", formula1='"interview,manual_review,supplement,reject"',
                            allow_blank=True)
    dv_rec.error = "请选择: interview / manual_review / supplement / reject"
    rec_col_letter = get_column_letter(4)
    dv_rec.add(f"{rec_col_letter}{header_row + 1}:{rec_col_letter}{header_row + nrows}")
    ws.add_data_validation(dv_rec)

    # 数据验证：人工复核结论列
    dv_review = DataValidation(type="list",
                               formula1='"待复核,进入面试,补充材料,不建议推进（人工确认）"',
                               allow_blank=True)
    review_col_letter = get_column_letter(13)
    dv_review.add(f"{review_col_letter}{header_row + 1}:{review_col_letter}{header_row + nrows}")
    ws.add_data_validation(dv_review)

    # 冻结
    ws.freeze_panes = ws.cell(header_row + 1, 1)

    wb.save(output_path)
    print(f"Done. {len(rows)} rows -> {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python flatten_batch_output.py <dify下载的csv> [输出文件名]")
        sys.exit(1)
    flatten(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
