from datetime import datetime
from io import BytesIO
from math import ceil

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import BatchTask, ItemStatus


HEADERS = [
    ("文件名", 24),
    ("候选人", 15),
    ("综合匹配分", 12),
    ("建议", 16),
    ("技能匹配", 12),
    ("相关经验", 12),
    ("项目相关性", 13),
    ("综合质量", 12),
    ("匹配技能", 26),
    ("待补充信息", 30),
    ("风险标记", 30),
    ("面试问题", 34),
    ("证据摘要", 40),
    ("人工复核结论", 24),
    ("备注", 28),
]


def _join(values: list[str]) -> str:
    return "\n".join(f"• {value}" for value in values)


def _estimated_lines(value: str, chars_per_line: int) -> int:
    if not value:
        return 1
    return sum(max(1, ceil(len(line) / chars_per_line)) for line in value.splitlines())


def build_workbook(task: BatchTask) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "候选人汇总"
    ws.sheet_view.showGridLines = False

    teal = PatternFill("solid", fgColor="0F766E")
    header_fill = PatternFill("solid", fgColor="D1FAE5")
    header_font = Font(bold=True, color="134E4A")
    thin = Side(style="thin", color="D1D5DB")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title = ws.cell(1, 1, "AI 简历初筛与岗位匹配助手｜人工复核汇总")
    title.fill = teal
    title.font = Font(bold=True, color="FFFFFF", size=16)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    for col, (name, width) in enumerate(HEADERS, 1):
        cell = ws.cell(3, col, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="0F766E"))
        ws.column_dimensions[get_column_letter(col)].width = width

    successful = [item for item in task.items if item.status == ItemStatus.success and item.result]
    for row_number, item in enumerate(successful, 4):
        result = item.result
        evidence = []
        for label, values in (
            ("技能", result.evidence.skill_match),
            ("经验", result.evidence.experience_relevance),
            ("项目", result.evidence.project_relevance),
            ("质量", result.evidence.overall_quality),
        ):
            if values:
                evidence.append(f"{label}：" + "；".join(values))
        values = [
            item.file_name,
            result.candidate_name,
            result.match_score,
            result.recommendation,
            result.dimension_scores.skill_match,
            result.dimension_scores.experience_relevance,
            result.dimension_scores.project_relevance,
            result.dimension_scores.overall_quality,
            _join(result.matched_skills),
            _join(result.missing_information),
            _join(result.risk_flags),
            _join(result.recommended_interview_questions),
            "\n".join(evidence),
            "",
            "",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_number, col, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        content_lines = max(
            _estimated_lines(str(values[8]), 24),
            _estimated_lines(str(values[9]), 26),
            _estimated_lines(str(values[10]), 26),
            _estimated_lines(str(values[11]), 30),
            _estimated_lines(str(values[12]), 36),
        )
        ws.row_dimensions[row_number].height = min(180, max(72, 17 * content_lines))

    end_row = max(4, len(successful) + 3)
    ws.auto_filter.ref = f"A3:O{end_row}"
    ws.freeze_panes = "A4"
    score_range = f"C4:C{end_row}"
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(
            operator="lessThan",
            formula=["60"],
            fill=PatternFill("solid", fgColor="FECACA"),
            font=Font(color="991B1B", bold=True),
        ),
    )
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(
            operator="between",
            formula=["60", "79"],
            fill=PatternFill("solid", fgColor="FEF3C7"),
            font=Font(color="92400E", bold=True),
        ),
    )
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["80"],
            fill=PatternFill("solid", fgColor="BBF7D0"),
            font=Font(color="166534", bold=True),
        ),
    )
    review_validation = DataValidation(
        type="list",
        formula1='"待复核,进入面试,补充材料,不建议推进（人工确认）"',
        allow_blank=True,
    )
    review_validation.add(f"N4:N{end_row}")
    ws.add_data_validation(review_validation)

    notes = wb.create_sheet("处理说明")
    notes.sheet_view.showGridLines = False
    notes.append(["项目", "内容"])
    notes.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    notes.append(["岗位 JD", task.job_description])
    notes.append(["补充评价要求", task.custom_instructions or "未填写"])
    notes.append(["成功数量", len(successful)])
    notes.append(["失败/取消数量", len(task.items) - len(successful)])
    notes.append([])
    notes.append(["文件名", "状态", "失败原因"])
    for item in task.items:
        if item.status != ItemStatus.success:
            notes.append([item.file_name, item.status.value, item.error or ""])
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 18
    notes.column_dimensions["C"].width = 72
    notes["A1"].fill = teal
    notes["B1"].fill = teal
    notes["A1"].font = notes["B1"].font = Font(bold=True, color="FFFFFF")
    notes.freeze_panes = "A2"
    for row in notes.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
