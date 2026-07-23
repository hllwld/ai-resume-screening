from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.export import build_workbook
from app.models import (
    BatchItem,
    BatchTask,
    DimensionScores,
    EvaluationResult,
    ItemStatus,
)


def test_export_has_summary_and_notes():
    result = EvaluationResult(
        candidate_name="候选人A",
        match_score=75,
        dimension_scores=DimensionScores(
            skill_match=80,
            experience_relevance=70,
            project_relevance=80,
            overall_quality=65,
        ),
        recommendation="manual_review",
        matched_skills=["Python", "RAG"],
    )
    now = datetime.now(UTC).isoformat()
    task = BatchTask(
        task_id="task-1",
        owner_id="owner-1",
        job_description="这是一个用于测试的岗位说明，长度满足模型要求。",
        items=[
            BatchItem(
                item_id="item-1",
                client_id="client-1",
                file_name="candidate.pdf",
                status=ItemStatus.success,
                result=result,
            )
        ],
        created_at=now,
        updated_at=now,
    )
    workbook = load_workbook(BytesIO(build_workbook(task)))
    assert workbook.sheetnames == ["候选人汇总", "处理说明"]
    assert workbook["候选人汇总"]["B4"].value == "候选人A"
    assert workbook["候选人汇总"]["C4"].value == 75
    assert len(workbook["候选人汇总"].conditional_formatting) == 1
    assert len(list(workbook["候选人汇总"].data_validations.dataValidation)) == 1
